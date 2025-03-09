import streamlit as st
import pandas as pd
import plotly.express as px
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import openpyxl

# 1️⃣ Carica il file Excel con i collegamenti ipertestuali
file_path_goes = "sfl_2005.xlsx"  # File sfl dal 2005 pagina GOES archivio
wb = openpyxl.load_workbook(file_path_goes, data_only=False)
ws = wb.active

# 2️⃣ Estrarre i link ipertestuali dalla colonna 'Snapshot Time'
links = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.hyperlink:
            links.append(cell.hyperlink.target)
        else:
            links.append(None)

# 3️⃣ Creare il DataFrame per i link
snapshot_times = [cell.value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1) for cell in row]
df_links = pd.DataFrame({"Snapshot Time": snapshot_times, "Real Link": links}).dropna(subset=["Real Link"])

# Funzione per contare le righe nella tabella 'Event#'
def get_event_count(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        tables = soup.find_all("table")
        for table in tables:
            headers = [th.text.strip() for th in table.find_all("th")]
            if "Event#" in headers:
                return len(table.find_all("tr")) - 1
        return 0
    except Exception:
        return 0

# 4️⃣ Applica la funzione di conteggio eventi
df_links["Event Count"] = df_links["Real Link"].apply(get_event_count)
df_links["Year"] = pd.to_datetime(df_links["Snapshot Time"], errors="coerce").dt.year
event_counts_links = df_links.groupby("Year")["Event Count"].sum()

# 5️⃣ Caricare i dati GOES
df_goes = pd.read_excel(file_path_goes)
df_goes.columns = df_goes.columns.str.strip()
df_goes["Snapshot Time"] = pd.to_datetime(df_goes["Snapshot Time"], errors="coerce")
df_goes = df_goes[df_goes["Snapshot Time"].dt.year >= 2005]
df_goes["Class"] = df_goes["Largest Event"].astype(str).str[0]
df_grouped_goes = df_goes.groupby([df_goes["Snapshot Time"].dt.year, "Class"]).size().reset_index(name="Count")

# 6️⃣ Caricare i dati FERMI
file_path_fermi = "GBM FERMI.xlsx"
df_fermi = pd.read_excel(file_path_fermi)
df_fermi.columns = df_fermi.columns.str.strip()
df_fermi["Date"] = pd.to_datetime(df_fermi["Date"], errors="coerce")
df_fermi = df_fermi[df_fermi["Date"].dt.year >= 2005]
df_grouped_fermi = df_fermi.groupby(df_fermi["Date"].dt.year).size().reset_index(name="Count")

# 🔹 INTERFACCIA STREAMLIT
st.sidebar.title("Navigazione")
menu = st.sidebar.selectbox("Seleziona una sezione:", ["🏠 Home", "📊 Dati GOES", "📈 Dati FERMI", "🔍 Dettagli Eventi"])

if menu == "📊 Dati GOES":
    st.title("📊 Dati GOES: Attività Solare")
    
    # Primo grafico (dati originali GOES)
    fig_goes = px.bar(df_grouped_goes, x="Snapshot Time", y="Count", color="Class",
                      title="Attività Solare nel Tempo (Dati GOES)",
                      labels={"Snapshot Time": "Anno", "Count": "Numero di Eventi"},
                      barmode="stack")
    st.plotly_chart(fig_goes)
    
    # Secondo grafico (eventi dai link estratti)
    st.write("### Eventi estratti dalle pagine web")
    fig, ax = plt.subplots(figsize=(10, 5))
    event_counts_links.plot(kind="bar", color="skyblue", edgecolor="black", ax=ax)
    ax.set_title("Numero Totale di Eventi per Anno (Dati dai Link)")
    ax.set_xlabel("Anno")
    ax.set_ylabel("Numero di Eventi")
    st.pyplot(fig)
    
    st.write("📊 **Dati elaborati (GOES):**")
    st.dataframe(df_goes)

if menu == "📈 Dati FERMI":
    st.title("📈 Dati FERMI: Attività Solare")
    fig_fermi = px.bar(df_grouped_fermi, x="Date", y="Count",
                       title="Numero Totale di Eventi Solari (Dati FERMI)",
                       labels={"Date": "Anno", "Count": "Numero di Eventi"})
    st.plotly_chart(fig_fermi)
    st.write("📊 **Dati elaborati (FERMI):**")
    st.dataframe(df_fermi)

if menu == "🏠 Home":
    st.title("Benvenuto nella Dashboard dell'Attività Solare")
    st.write("Questa dashboard visualizza l'attività solare nel tempo.")
    st.write("📊 Vai su 'Dati GOES' o 'Dati FERMI' per esplorare i dati.")

if menu == "🔍 Dettagli Eventi":
    st.title("🔍 Dettagli Eventi Solari")
    st.write("Approfondisci i dettagli degli eventi solari registrati.")
